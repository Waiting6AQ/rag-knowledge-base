"""
文件处理工具

提供上传保存、类型校验、文档加载器选择等功能。
"""
import hashlib
import os
from pathlib import Path
from fastapi import UploadFile, HTTPException
from langchain_community.document_loaders import TextLoader, PyPDFLoader, Docx2txtLoader
from langchain_core.documents import Document


class _ExcelLoader:
    """用 openpyxl 读取 Excel 的简易 Loader（兼容 LangChain loader 接口）"""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self) -> list[Document]:
        import openpyxl
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        docs = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if text.strip():
                    rows.append(text)
            if rows:
                docs.append(Document(
                    page_content="\n".join(rows),
                    metadata={"source": os.path.basename(self.file_path), "sheet": sheet_name},
                ))
        wb.close()
        return docs


# 扩展名 → 加载器映射
_LOADER_MAP = {
    ".txt": TextLoader,
    ".md": TextLoader,
    ".pdf": PyPDFLoader,
    ".docx": Docx2txtLoader,
    ".xlsx": _ExcelLoader,
}


def validate_extension(filename: str, allowed: list[str]) -> str:
    """校验文件扩展名是否合法，返回小写扩展名，不合法抛 HTTPException(400)"""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件类型 '{ext}'，允许：{', '.join(allowed)}",
        )
    return ext


def detect_loader(file_path: str, ext: str):
    """根据扩展名返回对应的 LangChain 文档加载器"""
    loader_cls = _LOADER_MAP.get(ext, TextLoader)
    if loader_cls is TextLoader:
        return loader_cls(file_path, encoding="utf-8")
    return loader_cls(file_path)


async def save_upload(file: UploadFile, target_dir: str, max_size_bytes: int) -> tuple[Path, str]:
    """流式保存文件，边写边算 SHA256、边校验大小，返回 (保存路径, 文件哈希)

    全程内存仅占缓冲区块大小，避免大文件一次性读入内存。
    """
    os.makedirs(target_dir, exist_ok=True)
    save_path = Path(target_dir) / file.filename
    if save_path.exists():
        # 同名文件已存在：加 (1) (2) 后缀，类似 Windows 行为
        stem = save_path.stem
        suffix = save_path.suffix
        counter = 1
        while True:
            candidate = Path(target_dir) / f"{stem} ({counter}){suffix}"
            if not candidate.exists():
                save_path = candidate
                break
            counter += 1

    sha256_hash = hashlib.sha256()
    total_size = 0
    with open(save_path, "wb") as f:
        while True:
            chunk = await file.read(1024 * 1024)   # 1MB 缓冲，平衡循环次数与内存占用
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > max_size_bytes:        # 实时拦截：超限立即清理并拒绝
                f.close()
                if os.path.exists(save_path):
                    os.remove(save_path)
                raise HTTPException(
                    status_code=413,
                    detail=f"文件超过大小限制 {max_size_bytes // (1024 * 1024)}MB，请压缩后重新上传",
                )
            sha256_hash.update(chunk)
            f.write(chunk)

    return save_path, sha256_hash.hexdigest()
